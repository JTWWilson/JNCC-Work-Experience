from openpyxl import load_workbook
import yaml

OUTPUT_FILE_NAME = 'new_config.yaml'

PROFORMA_PATH = r'Image_proforma_GD_standard_v2.xlsx'

proforma = load_workbook(PROFORMA_PATH, data_only=True)['Fields_Descriptions']


def iter_rows_until_blank(sheet, min_row=0, max_row=float('inf')):
    """
    Iterates through a spreadsheet's rows until it hits a row with a blank A column value
    :param sheet: Spreadsheet
    :param min_row: Row to start iterating from (inclusive)
    :param max_row: Row to end iterating at (inclusive)
    :yield: Each row
    """
    count = min_row - 1
    values = list(sheet.values)
    while count <= max_row:
        try:
            current_row = values[count]
        except IndexError:
            raise StopIteration
        if current_row[0] is None:
            raise StopIteration
        yield current_row
        count += 1

config = {'File_Name': None,
          'Species_of_Note_Workbook': None,
          'Analysis_Workbook':
              {'Path': None,
               'Sheet_Name': None,
               'Field_Row': None,
               'Data_Starts_at_Row_Number': None},
          'Species_Matrix':
              {'Path': None,
               'Sheet_Name': None,
               'Species_Name_Column_Number': None,
               'Blank_Entry_Looks_Like': None},
          'Fields': {}
          }

# Iterates through proforma adding the field names to a new config
for row in iter_rows_until_blank(proforma, 3):
    config['Fields'][row[0]] = [[], ""]


with open(OUTPUT_FILE_NAME, 'w') as outfile:
    yaml.dump(config, outfile, default_flow_style=False)
